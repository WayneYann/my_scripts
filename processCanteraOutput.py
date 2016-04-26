"""
Save data from a bunch of Cantera simulations in Excel.
"""
import sys
import os
from openpyxl import Workbook, load_workbook
import re
#################################################################################
#/Users/belinda/Code/Cantera/Onishuck_thermal_decomp/more_HBI/y0_high/output.txt
wb = Workbook()
#ws = wb.create_sheet(title="Full Mech Data")
#ws.title = "Data"
#n = 0
for root, dirs, files in os.walk("./Onishuck_thermal_decomp/more_HBI/feb2016"):
    if "output_all_std.txt" in files:
        print os.path.split(os.path.split(root)[-2])[-1] + '_' + os.path.split(root)[-1]
        ws = wb.create_sheet(title=os.path.split(os.path.split(root)[-2])[-1] + '_' + os.path.split(root)[-1])
        #n += 1
        with open(os.path.join(root, "output_all_std.txt"), 'r') as cantera_output:
           # ws.cell(row=1, column=(3*n-2)).value = os.path.split(root)[-1]
            m = 1
            for line in cantera_output:
                line = re.sub('[\[\]]', '', line)
                m += 1
                i = 0
                for data in line.split():
                    i += 1
                    ws.cell(row = m, column = i).value = data
                    ws.cell(row = m, column = i).data_type = 'n'
                    ws.cell(row = m, column = i).number_format = '0.00E+00'
               # elif len(line.split()) == 4:
               #     i = 5
               #     for data in line.split():
               #         i += 1
               #         ws.cell(row = m, column = i).value = data
               #         ws.cell(row = m, column = i).data_type = 'n'
               #         ws.cell(row = m, column = i).number_format = '0.00E+00'
               # else:
               #     i = 9
               #     ws.cell(row = m, column = 10).value = line.strip()
               #     ws.cell(row = m, column = 10).data_type = 'n'
               #     ws.cell(row = m, column = 10).number_format = '0.00E+00'
                #ws.cell(row = m, column=(3*n-2)).value = time
                #ws.cell(row = m, column=(3*n-2)).data_type = 'n'
                #ws.cell(row = m, column=(3*n-2)).number_format = "0.00E+00"
                #ws.cell(row = m, column=(3*n-1)).value = conc
                #ws.cell(row = m, column=(3*n-1)).data_type = 'n'
                #ws.cell(row = m, column=(3*n-1)).number_format = "0.00E+00"
wb.save(filename = "./Onishuck_thermal_decomp/more_HBI/feb2016/result.xlsx")
